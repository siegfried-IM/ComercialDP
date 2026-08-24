# -*- coding: utf-8 -*-
"""Cruza las presentaciones de IQVIA contra la lista de precios y calcula un
PSL PONDERADO por mercado.

El tablero razona por mercado; la lista tiene un PSL por presentación y dentro de
un mismo mercado los precios llegan a diferir 45x. El ponderador es el mix real de
unidades (datos/presentaciones.json, MAT de Siegfried).

El match se hace por concentración + tamaño de envase, que es lo único común entre
"ACEMUK TABL EFERV 600MG x 10" y ("ACEMUK 600 mg", "Comprimidos", "con 10").
Lo que importa no es cuántas filas matchean sino qué FRACCIÓN DE LAS UNIDADES
queda cubierta: una presentación chica sin precio no mueve el resultado, la que
más vende sí. Por eso el reporte es en unidades.

Uso:  python cruzar_precios.py <lista.xlsx>
Salida: ../datos/psl_ponderado.json
"""
import json, os, re, sys, unicodedata
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
DATA = os.path.join(HERE, "..", "datos")
COL_PROD, COL_PRES, COL_ENV, COL_PSL = 7, 8, 9, 11


def norm(s):
    """El % se conserva: es lo único que separa MICROSONA 0,5% de 1% y de 2%,
    que comparten marca, forma y envase."""
    s = unicodedata.normalize("NFKD", str(s or "")).encode("ascii", "ignore").decode().upper()
    return re.sub(r"\s+", " ", re.sub(r"[^A-Z0-9.,% ]", " ", s)).strip()


def contenido(txt):
    """Contenido del envase unitario: '15G', '60ML', 'con 15 gr', 'con 20 ml'.

    Cremas, jarabes y gotas vienen en IQVIA como 'x 1' (un pomo, un frasco) y en
    la lista como 'con 15 gr': el conteo de envases no cruza porque no miden lo
    mismo. Para esas formas el contenido ES la clave.
    """
    # El contenido es el token que precede al 'x N'. Ni el primero ni el ultimo
    # sirven: en 'PALDAR CREMA 0.30G 15G x 1' el primero es la dosis, y en
    # 'ACEMUK GRAN 2.00G 100ML x 1 /5ML' el ultimo es la dosis por 5 ml.
    t = norm(txt)
    # se corta solo en la "x" de IQVIA: la lista escribe "con 15 gr" y cortar
    # ahi dejaba el texto vacio
    corte = re.search(r"\bX\s*\d", t)
    if corte:
        t = t[:corte.start()]
    ms = list(re.finditer(r"(\d+[.,]?\d*)\s*(GR|G|ML)\b", t))
    if not ms:
        return None
    m = ms[-1]
    v = float(m.group(1).replace(",", "."))
    return (round(v, 3), "ML" if m.group(2) == "ML" else "G")


def concentracion(txt):
    """Dosis del texto, como {(valor, unidad)}. '600MG' -> {(600,'MG')}.

    Se excluye el contenido del envase: en 'MICROSONA CREMA 1% 15G x 1' el 15G es
    el tamaño del pomo, no la dosis, y tomarlo como 15.000 mg inventa un conflicto
    de concentración que descarta el match correcto.
    """
    t = norm(txt)
    cont = contenido(t)
    out = set()
    for m in re.finditer(r"(\d+[.,]?\d*)\s*(MG|G|MCG|UI|ML|%)", t):
        v = float(m.group(1).replace(",", "."))
        u = m.group(2)
        if u == "G":
            v, u = v * 1000, "MG"        # 1 g = 1000 mg, para que 2.00G cruce con 2000MG
        if cont and ((u == "MG" and abs(v - cont[0] * 1000) < 1e-6 and cont[1] == "G")
                     or (u == "ML" and abs(v - cont[0]) < 1e-6 and cont[1] == "ML")):
            continue                      # es el contenido, no la dosis
        out.add((round(v, 3), u))
    return out


def concuerda_conc(a, b):
    """¿Son la misma dosis? Acepta que IQVIA desglose los componentes y la lista
    publique la suma: Entresto sale como '24MG/26mg' y se lista como '50 mg'."""
    if not a or not b:
        return None                       # un lado no la declara: no decide
    if a & b:
        return True
    for x, y in ((a, b), (b, a)):
        mg = [v for v, u in x if u == "MG"]
        if len(mg) > 1 and any(u == "MG" and abs(sum(mg) - v) < 0.51 for v, u in y):
            return True
    return False


def envase(txt):
    """Cantidad de envases: 'x 10', 'con 10', 'X30'.

    El \\b delante es imprescindible: sin el, la X final de marcas como EMPAX
    hacia leer 'EMPAX 10 MG' como envase 10 en vez del 30 real, y el producto
    entero quedaba sin precio.
    """
    m = re.search(r"\b(?:X|CON)\s*(\d+)\b", norm(txt))
    return int(m.group(1)) if m else None


def es_multipack(env_txt):
    """Pack de distribucion (dispenser, pack de N cajas), no un envase de venta.

    IQVIA cuenta cajas individuales; estas filas son N cajas juntas y su precio es
    el del bulto. Mezclarlas hacia leer 'ACEMUK 600 x 20' a $152.621 cuando la caja
    de 20 comprimidos vale $16.050: un error de 9x en el producto de mayor volumen.
    """
    t = norm(env_txt)
    return ("PACK" in t) or ("DISPENSER" in t)


def envase_lista(env_txt):
    """Envase de la lista: manda el 'con N'. En 'Pack x 20 con 5 tab, cada uno' el
    20 son los tubos del bulto y el 5 es lo que trae cada caja."""
    m = re.search(r"\bCON\s*(\d+)\b", norm(env_txt))
    if m:
        return int(m.group(1))
    return envase(env_txt)


def marca(txt):
    return norm(txt).split(" ")[0] if norm(txt) else ""


def main():
    xlsx = sys.argv[1]
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    ws = wb.active
    lista = []
    for r in range(16, ws.max_row + 1):
        p, psl = ws.cell(r, COL_PROD).value, ws.cell(r, COL_PSL).value
        if not p or not isinstance(psl, (int, float)):
            continue
        pres_c, env_c = ws.cell(r, COL_PRES).value or "", ws.cell(r, COL_ENV).value or ""
        txt = "%s %s %s" % (p, pres_c, env_c)
        if es_multipack(env_c):
            continue                      # bulto de distribucion: IQVIA cuenta cajas
        lista.append({"prod": str(p).strip(), "txt": txt, "psl": float(psl),
                      "marca": marca(p),
                      # la concentración puede estar en el nombre o en la presentación
                      "conc": concentracion(p) | concentracion(pres_c) | concentracion(env_c),
                      # el envase sale SÓLO de su columna: leerlo del nombre mezcla la marca
                      "env": envase_lista(env_c),
                      "cont": contenido(env_c),
                      "toks": set(norm(p).split())})

    pres = json.load(open(os.path.join(DATA, "presentaciones.json"), encoding="utf-8"))
    salida, resumen = {}, []
    for prod, mix in pres["datos"].items():
        mprod = marca(prod)
        tot_u = sum(mix.values())
        num = cubierto = 0.0
        detalle, sin = [], []
        for nombre, u in mix.items():
            c, e, cont = concentracion(nombre), envase(nombre), contenido(nombre)
            toks = set(norm(nombre).split())
            # Los candidatos salen de la marca de la PRESENTACION, no de la del
            # mercado: Siegfried tiene marcas propias dentro de un mismo mercado
            # (FER-IN-SOL en Siderblut, CEFACAR en Cefalexiona, MICORAL e IKOLAN en
            # Micomazol, ALERGICAL en Alidial, CITRAMAR en Calcio Base). Filtrando
            # por la marca del mercado quedaban valorizadas al precio de la marca
            # principal, que es otro producto y otro precio.
            mpres = marca(nombre)
            cand = [x for x in lista if x["marca"].startswith(mpres[:5])]
            if not cand:
                cand = [x for x in lista if x["marca"].startswith(mprod[:5])]
            best, score = None, -1
            for x in cand:
                # Condición dura: o coincide el conteo de envases, o —para las
                # formas unitarias— coincide el contenido en gramos/ml.
                por_conteo = (e is not None and x["env"] == e and not x["cont"])
                por_contenido = (cont is not None and x["cont"] == cont)
                # Pack multiple: IQVIA trae 'ROACCUTAN 30+30 20mg x 60' y la lista
                # solo publica el de 30. Un envase de 60 son dos de 30, asi que el
                # precio se prorratea. Solo hacia arriba (k>=2): suponer que un
                # envase chico vale una fraccion del grande seria falso, porque el
                # precio por unidad sube al achicar el envase.
                mult = 1
                if not (por_conteo or por_contenido) and e and x["env"] and not x["cont"]:
                    if e > x["env"] and e % x["env"] == 0:
                        mult = e // x["env"]
                        por_conteo = True
                if not (por_conteo or por_contenido):
                    continue
                s = 3 - (0.5 if mult > 1 else 0)   # el match exacto gana al prorrateado
                ok = concuerda_conc(x["conc"], c)
                if ok is True:
                    s += 2
                elif ok is False:
                    s = -99               # dosis distinta => es otra presentación
                # ok None: un lado no la declara ("SIDERBLUT" a secas), no penaliza
                s += len(x["toks"] & toks) * 0.1     # desempate por nombre
                if s > score:
                    score, best = s, x
            if best and score >= 2.5:        # envase compatible y sin conflicto de dosis
                k = (e // best["env"]) if (best["env"] and e and e > best["env"]
                                           and e % best["env"] == 0 and not best["cont"]) else 1
                psl = best["psl"] * k
                num += u * psl; cubierto += u
                detalle.append({"iqvia": nombre, "lista": best["prod"], "env": best["env"],
                                "psl": psl, "u": u, "packs": k})
            else:
                sin.append({"iqvia": nombre, "u": u})
        psl_pond = (num / cubierto) if cubierto else None
        salida[prod] = {"psl_ponderado": round(psl_pond, 2) if psl_pond else None,
                        "unidades_mat": round(tot_u, 1),
                        "cobertura": round(cubierto / tot_u, 4) if tot_u else 0,
                        "match": detalle, "sin_precio": sin}
        resumen.append((cubierto / tot_u if tot_u else 0, prod, psl_pond, tot_u, len(sin)))

    with open(os.path.join(DATA, "psl_ponderado.json"), "w", encoding="utf-8") as f:
        json.dump({"meta": {"lista": os.path.basename(xlsx), "ventana": pres["meta"]["ventana"],
                            "hasta": pres["meta"]["label"],
                            "precio": "PRECIO DE VENTA DROGUERIA SIN IVA"},
                   "datos": salida}, f, ensure_ascii=False, indent=1)

    resumen.sort()
    print("%-14s %9s %13s %12s  %s" % ("producto", "cobert.", "PSL ponder.", "unid. MAT", "sin precio"))
    for cob, prod, psl, tot, nsin in resumen:
        print("%-14s %8.1f%% %13s %12s  %s" % (
            prod, cob * 100, "{:,.0f}".format(psl) if psl else "—", "{:,.0f}".format(tot),
            nsin or ""))
    tu = sum(r[3] for r in resumen)
    tc = sum(r[3] * r[0] for r in resumen)
    print()
    print("cobertura global sobre unidades: %.2f%%  (%s de %s u. MAT)" % (
        100 * tc / tu, "{:,.0f}".format(tc), "{:,.0f}".format(tu)))


if __name__ == "__main__":
    main()
